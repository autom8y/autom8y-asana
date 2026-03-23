"""Tests for ExcelFormatEngine and compose_excel.

Per TDD-data-attachment-bridge-platform Section 7.
Unit tests for the Excel format engine including workbook structure,
sheet content, protocol conformance, and edge cases.
"""

from __future__ import annotations

import io
from typing import Any

import pytest
from openpyxl import load_workbook

from autom8_asana.automation.workflows.payment_reconciliation.formatter import (
    ExcelFormatEngine,
    compose_excel,
)
from autom8_asana.automation.workflows.protocols import FormatEngine

# --- Helpers ---


def _sample_rows() -> list[dict[str, Any]]:
    """Standard sample rows for tests."""
    return [
        {
            "date": "2026-01-01",
            "spend": 100.0,
            "payments": 50.0,
            "office_phone": "+1234567890",
        },
        {
            "date": "2026-01-15",
            "spend": 200.0,
            "payments": 100.0,
            "office_phone": "+1234567890",
        },
        {
            "date": "2026-02-01",
            "spend": 150.0,
            "payments": 75.0,
            "office_phone": "+1234567890",
        },
    ]


def _sample_data(rows: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    """Standard data dict for tests."""
    return {
        "rows": _sample_rows() if rows is None else rows,
        "office_phone": "***-***-7890",
        "vertical": "chiropractic",
        "business_name": "Test Business",
    }


def _load_workbook(excel_bytes: bytes) -> Any:
    """Load an openpyxl Workbook from bytes."""
    return load_workbook(io.BytesIO(excel_bytes))


# ---------------------------------------------------------------------------
# TestExcelFormatEngineAttributes
# ---------------------------------------------------------------------------


class TestExcelFormatEngineAttributes:
    """Verify engine class attributes."""

    def test_content_type(self) -> None:
        engine = ExcelFormatEngine()
        assert engine.content_type == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_file_extension(self) -> None:
        engine = ExcelFormatEngine()
        assert engine.file_extension == ".xlsx"


# ---------------------------------------------------------------------------
# TestRenderProducesValidXlsx
# ---------------------------------------------------------------------------


class TestRenderProducesValidXlsx:
    """Test that render() produces valid .xlsx content."""

    def test_render_produces_valid_xlsx(self) -> None:
        """render() returns bytes that openpyxl can open without error."""
        engine = ExcelFormatEngine()
        result = engine.render(_sample_data())
        assert isinstance(result, bytes)
        assert len(result) > 0

        # Should be loadable by openpyxl
        wb = _load_workbook(result)
        assert wb is not None
        wb.close()


# ---------------------------------------------------------------------------
# TestRenderSummarySheet
# ---------------------------------------------------------------------------


class TestRenderSummarySheet:
    """Test Summary sheet content."""

    def test_render_summary_sheet(self) -> None:
        """Workbook has a 'Summary' sheet at index 0 with business name and vertical."""
        engine = ExcelFormatEngine()
        result = engine.render(_sample_data())
        wb = _load_workbook(result)

        assert wb.sheetnames[0] == "Summary"
        ws = wb["Summary"]

        # Check business name and vertical are present
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "Test Business" in values
        assert "chiropractic" in values
        wb.close()

    def test_render_summary_aggregate_totals(self) -> None:
        """Summary sheet contains aggregate totals."""
        engine = ExcelFormatEngine()
        result = engine.render(_sample_data())
        wb = _load_workbook(result)

        ws = wb["Summary"]
        # Find the row with "Total Rows"
        found_total_rows = False
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Total Rows":
                assert ws.cell(row=row, column=2).value == 3
                found_total_rows = True
                break
        assert found_total_rows
        wb.close()

    def test_render_summary_masked_phone(self) -> None:
        """Summary sheet shows masked phone, not raw."""
        engine = ExcelFormatEngine()
        data = _sample_data()
        data["office_phone"] = "***-***-7890"
        result = engine.render(data)
        wb = _load_workbook(result)

        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "***-***-7890" in values
        wb.close()


# ---------------------------------------------------------------------------
# TestRenderDetailSheet
# ---------------------------------------------------------------------------


class TestRenderDetailSheet:
    """Test Reconciliation detail sheet content."""

    def test_render_detail_sheet(self) -> None:
        """Workbook has a 'Reconciliation' sheet with header row and correct row count."""
        engine = ExcelFormatEngine()
        rows = _sample_rows()
        result = engine.render(_sample_data(rows))
        wb = _load_workbook(result)

        assert "Reconciliation" in wb.sheetnames
        ws = wb["Reconciliation"]

        # Header row + data rows
        assert ws.max_row == len(rows) + 1

        # Header matches keys of first row
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == list(rows[0].keys())
        wb.close()

    def test_render_detail_sheet_data_values(self) -> None:
        """Detail sheet contains actual row data."""
        engine = ExcelFormatEngine()
        rows = [{"date": "2026-03-01", "spend": 42.5, "payments": 21.0}]
        result = engine.render(_sample_data(rows))
        wb = _load_workbook(result)

        ws = wb["Reconciliation"]
        assert ws.cell(row=2, column=1).value == "2026-03-01"
        assert ws.cell(row=2, column=2).value == 42.5
        assert ws.cell(row=2, column=3).value == 21.0
        wb.close()


# ---------------------------------------------------------------------------
# TestRenderEmptyRows
# ---------------------------------------------------------------------------


class TestRenderEmptyRows:
    """Test rendering with empty data."""

    def test_render_empty_rows(self) -> None:
        """render with empty rows produces valid workbook."""
        engine = ExcelFormatEngine()
        data = _sample_data([])
        result = engine.render(data)
        wb = _load_workbook(result)

        # Summary should still exist
        assert "Summary" in wb.sheetnames
        # Detail sheet should have a placeholder message
        ws = wb["Reconciliation"]
        assert ws.cell(row=1, column=1).value == "No reconciliation data available"
        wb.close()


# ---------------------------------------------------------------------------
# TestRenderMultiPeriodSheets
# ---------------------------------------------------------------------------


class TestRenderMultiPeriodSheets:
    """Test per-period sheet grouping."""

    def test_render_multi_period_sheets(self) -> None:
        """When rows contain multiple periods, separate sheets are created."""
        rows = [
            {"date": "2026-01-01", "spend": 100.0, "period": "Jan-2026"},
            {"date": "2026-01-15", "spend": 200.0, "period": "Jan-2026"},
            {"date": "2026-02-01", "spend": 150.0, "period": "Feb-2026"},
        ]
        engine = ExcelFormatEngine()
        result = engine.render(_sample_data(rows))
        wb = _load_workbook(result)

        # Should have Summary, Reconciliation, Jan-2026, Feb-2026
        assert "Summary" in wb.sheetnames
        assert "Reconciliation" in wb.sheetnames
        assert "Jan-2026" in wb.sheetnames
        assert "Feb-2026" in wb.sheetnames

        # Verify Jan sheet has 2 data rows + header
        ws_jan = wb["Jan-2026"]
        assert ws_jan.max_row == 3  # 1 header + 2 data

        # Verify Feb sheet has 1 data row + header
        ws_feb = wb["Feb-2026"]
        assert ws_feb.max_row == 2  # 1 header + 1 data
        wb.close()

    def test_render_single_period_no_extra_sheets(self) -> None:
        """When all rows share a single period, no extra period sheets are created."""
        rows = [
            {"date": "2026-01-01", "spend": 100.0, "period": "Jan-2026"},
            {"date": "2026-01-15", "spend": 200.0, "period": "Jan-2026"},
        ]
        engine = ExcelFormatEngine()
        result = engine.render(_sample_data(rows))
        wb = _load_workbook(result)

        # Only Summary and Reconciliation -- no separate period sheet
        assert wb.sheetnames == ["Summary", "Reconciliation"]
        wb.close()


# ---------------------------------------------------------------------------
# TestComposeExcelConvenience
# ---------------------------------------------------------------------------


class TestComposeExcelConvenience:
    """Test the compose_excel convenience function."""

    def test_compose_excel_returns_tuple(self) -> None:
        """compose_excel returns (bytes, row_count) tuple."""
        rows = _sample_rows()
        excel_bytes, row_count = compose_excel(
            rows,
            office_phone="***-***-7890",
            vertical="chiropractic",
            business_name="Test",
        )
        assert isinstance(excel_bytes, bytes)
        assert row_count == len(rows)

    def test_compose_excel_row_count_matches_input(self) -> None:
        """Row count matches the number of input rows."""
        rows = [{"x": 1}, {"x": 2}, {"x": 3}, {"x": 4}, {"x": 5}]
        _, row_count = compose_excel(
            rows,
            office_phone="***-***-0000",
            vertical="dental",
        )
        assert row_count == 5


# ---------------------------------------------------------------------------
# TestFormatEngineProtocolConformance
# ---------------------------------------------------------------------------


class TestFormatEngineProtocolConformance:
    """Test structural conformance to FormatEngine protocol."""

    def test_format_engine_protocol_conformance(self) -> None:
        """ExcelFormatEngine is recognized as FormatEngine (runtime_checkable)."""
        engine = ExcelFormatEngine()
        assert isinstance(engine, FormatEngine)

    def test_format_engine_has_content_type(self) -> None:
        """FormatEngine protocol requires content_type."""
        engine = ExcelFormatEngine()
        assert hasattr(engine, "content_type")
        assert isinstance(engine.content_type, str)

    def test_format_engine_has_file_extension(self) -> None:
        """FormatEngine protocol requires file_extension."""
        engine = ExcelFormatEngine()
        assert hasattr(engine, "file_extension")
        assert isinstance(engine.file_extension, str)

    def test_format_engine_has_render(self) -> None:
        """FormatEngine protocol requires render(data) -> bytes."""
        engine = ExcelFormatEngine()
        assert callable(engine.render)
