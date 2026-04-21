"""Excel format engine for payment reconciliation workbooks.

Per ADR-bridge-format-engine: Implements FormatEngine protocol.
Per TDD-data-attachment-bridge-platform Section 7.

Produces multi-sheet .xlsx workbooks with Summary and Reconciliation
detail sheets, plus optional per-period grouping.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from autom8y_api_schemas import OfficePhoneField  # noqa: TC002
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExcelFormatEngine:
    """Format engine producing .xlsx reconciliation workbooks.

    Per ADR-bridge-format-engine: Implements FormatEngine protocol.
    Per TDD-data-attachment-bridge-platform Section 7.

    Attributes:
        content_type: MIME type for .xlsx files.
        file_extension: File extension including dot.
    """

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    file_extension = ".xlsx"

    def render(self, data: dict[str, Any]) -> bytes:
        """Transform reconciliation data dict into Excel workbook bytes.

        Args:
            data: Dict with keys:
                - "rows" (list[dict]): Reconciliation data rows.
                - "office_phone" (str): Masked phone for header display.
                - "vertical" (str): Business vertical.
                - "business_name" (str | None): Business name for title.

        Returns:
            .xlsx file content as bytes.
        """
        rows: list[dict[str, Any]] = data.get("rows", [])
        office_phone: OfficePhoneField = data.get("office_phone", "")
        vertical: str = data.get("vertical", "")
        business_name: str | None = data.get("business_name")

        wb = Workbook()

        # --- Summary sheet (tab 0) ---
        ws_summary = wb.active
        assert ws_summary is not None
        ws_summary.title = "Summary"

        ws_summary.append(["Payment Reconciliation Report"])
        ws_summary.append([])
        ws_summary.append(["Business Name", business_name or "Unknown"])
        ws_summary.append(["Vertical", vertical])
        ws_summary.append(["Phone", office_phone])
        ws_summary.append(["Generated", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")])
        ws_summary.append([])

        # Aggregate totals from row data
        total_rows = len(rows)
        total_spend = _safe_sum(rows, "spend")
        total_payments = _safe_sum(rows, "payments")

        ws_summary.append(["Total Rows", total_rows])
        ws_summary.append(["Total Spend", total_spend])
        ws_summary.append(["Total Payments", total_payments])

        # --- Detail sheet (tab 1) ---
        ws_detail = wb.create_sheet("Reconciliation")

        if rows:
            headers = list(rows[0].keys())
            ws_detail.append(headers)
            for row in rows:
                ws_detail.append([row.get(h) for h in headers])
            _auto_width(ws_detail, headers)
        else:
            ws_detail.append(["No reconciliation data available"])

        # --- Optional per-period sheets ---
        if rows and any("period" in r for r in rows):
            periods: dict[str, list[dict[str, Any]]] = {}
            for row in rows:
                period_key = row.get("period")
                if period_key is not None:
                    periods.setdefault(str(period_key), []).append(row)

            # Only create per-period sheets if multiple periods exist
            if len(periods) > 1:
                for period_name, period_rows in sorted(periods.items()):
                    # Sanitize sheet name (max 31 chars, no invalid chars)
                    safe_name = _sanitize_sheet_name(period_name)
                    ws_period = wb.create_sheet(safe_name)
                    headers = list(period_rows[0].keys())
                    ws_period.append(headers)
                    for row in period_rows:
                        ws_period.append([row.get(h) for h in headers])
                    _auto_width(ws_period, headers)

        # Serialize to bytes
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def compose_excel(
    response_data: list[dict[str, Any]],
    *,
    office_phone: str,
    vertical: str,
    business_name: str | None = None,
) -> tuple[bytes, int]:
    """Compose Excel workbook from reconciliation data.

    Args:
        response_data: Rows from InsightsResponse.data.
        office_phone: Masked phone number for header display.
        vertical: Business vertical.
        business_name: Business name for workbook title.

    Returns:
        Tuple of (excel_bytes, row_count).
    """
    engine = ExcelFormatEngine()
    excel_bytes = engine.render(
        {
            "rows": response_data,
            "office_phone": office_phone,
            "vertical": vertical,
            "business_name": business_name,
        }
    )
    return excel_bytes, len(response_data)


# --- Private helpers ---


def _safe_sum(rows: list[dict[str, Any]], key: str) -> float:
    """Sum a numeric field across rows, ignoring missing/non-numeric values."""
    total = 0.0
    for row in rows:
        val = row.get(key)
        if isinstance(val, int | float):
            total += val
    return total


def _auto_width(ws: Any, headers: list[str]) -> None:
    """Set column widths based on header length (heuristic)."""
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        # Min width 10, max width 40, based on header length + padding
        width = min(max(len(str(header)) + 4, 10), 40)
        ws.column_dimensions[col_letter].width = width


def _sanitize_sheet_name(name: str) -> str:
    """Sanitize a string for use as an Excel sheet name.

    Excel sheet names must be <= 31 characters and cannot contain
    certain special characters.
    """
    # Remove invalid characters
    invalid_chars = r"[]:*?/\\"
    sanitized = name
    for ch in invalid_chars:
        sanitized = sanitized.replace(ch, "_")
    return sanitized[:31] or "Sheet"
